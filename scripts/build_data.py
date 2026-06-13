from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl


SOURCE = Path("/Users/nerolifu/Downloads/Summary of software licensing (1).xlsx")
OUT = Path(__file__).resolve().parents[1] / "data" / "dashboard-data.json"


DETAIL_SHEETS = {
    "MY Details Breakdown": {
        "country": "Malaysia",
        "countryCode": "MY",
        "currency": "RM",
        "amountHeader": "Amount (RM) ",
        "totalHeader": "Total (RM)",
    },
    "HK Details Breakdown": {
        "country": "Hong Kong",
        "countryCode": "HK",
        "currency": "HKD",
        "amountHeader": "Amount (HKD) ",
        "totalHeader": "Total (HKD)",
    },
}

PROJECT_ALIASES = {
    "vio": "VIO + Mini APP",
    "vio 4": "VIO + Mini APP",
    "miniapp": "VIO + Mini APP",
    "mini app": "VIO + Mini APP",
    "vouchain": "VouChain",
    "hr systems": "HR Systems",
    "mcp service": "MCP Service",
}


def clean(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return re.sub(r"\s+", " ", value).strip()
    return value


def month_key(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, str) and value.strip():
        return value.strip()[:7]
    return ""


def amount(value):
    if value is None or value == "":
        return 0.0
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return 0.0


def normalize_project(value):
    value = clean(value)
    if not value:
        return "Unassigned"
    key = value.casefold()
    return PROJECT_ALIASES.get(key, value)


def split_projects(value):
    raw = clean(value)
    if not raw:
        return ["Unassigned"]
    normalized = []
    for part in re.split(r"\s*[/,]\s*", raw):
        project = normalize_project(part)
        if project and project not in normalized:
            normalized.append(project)
    return normalized or ["Unassigned"]


def rows_from_sheet(workbook, sheet_name, config):
    ws = workbook[sheet_name]
    headers = [clean(cell.value) for cell in ws[3]]
    index = {header: i for i, header in enumerate(headers)}

    rows = []
    for excel_row in ws.iter_rows(min_row=4, values_only=True):
        employee = clean(excel_row[index["Employee's Name"]])
        vendor = clean(excel_row[index["Vendor"]])
        plan = clean(excel_row[index["Subscription Plan"]])
        month = month_key(excel_row[index["Month"]])
        line_amount = amount(excel_row[index[config["amountHeader"].strip()]])
        total = amount(excel_row[index[config["totalHeader"].strip()]])
        projects = split_projects(excel_row[index["Project"]]) if "Project" in index else ["Unassigned"]

        if not employee or not vendor or not month or total <= 0:
            continue
        project_total = round(total / len(projects), 2)
        project_allocations = [
            {
                "project": project,
                "total": project_total if i < len(projects) - 1 else round(total - project_total * (len(projects) - 1), 2),
            }
            for i, project in enumerate(projects)
        ]

        rows.append(
            {
                "country": config["country"],
                "countryCode": config["countryCode"],
                "currency": config["currency"],
                "employee": employee,
                "vendor": vendor,
                "plan": plan or "Unspecified",
                "subscription": f"{vendor} - {plan}" if plan else vendor,
                "month": month,
                "amount": line_amount,
                "total": total,
                "projects": projects,
                "projectAllocations": project_allocations,
                "supportUser": clean(excel_row[index["Support User"]]),
                "remark": clean(excel_row[index["Remark"]]),
            }
        )
    return rows


def sum_by(rows, keys, amount_key="total"):
    grouped = defaultdict(float)
    for row in rows:
        grouped[tuple(row[key] for key in keys)] += row[amount_key]
    return [
        {**{key: group[i] for i, key in enumerate(keys)}, "total": round(total, 2)}
        for group, total in grouped.items()
    ]


def main():
    wb = openpyxl.load_workbook(SOURCE, data_only=True, read_only=False)
    rows = []
    for sheet, config in DETAIL_SHEETS.items():
        rows.extend(rows_from_sheet(wb, sheet, config))

    rows.sort(key=lambda row: (row["month"], row["countryCode"], row["employee"], row["vendor"]))

    payload = {
        "sourceFile": str(SOURCE),
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "rows": rows,
        "monthlyByCountry": sorted(sum_by(rows, ["countryCode", "country", "currency", "month"]), key=lambda row: (row["month"], row["countryCode"])),
        "employeeTotals": sorted(sum_by(rows, ["countryCode", "country", "currency", "employee"]), key=lambda row: row["total"], reverse=True),
        "subscriptionTotals": sorted(sum_by(rows, ["countryCode", "country", "currency", "subscription", "vendor", "plan"]), key=lambda row: row["total"], reverse=True),
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUT} with {len(rows)} rows")


if __name__ == "__main__":
    main()
